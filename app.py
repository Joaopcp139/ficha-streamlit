
import streamlit as st
import PyPDF2
import re
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook

def extrair_dados_pdf(pdf_file):
    pdf_reader = PyPDF2.PdfReader(pdf_file)
    texto = ""
    for page in pdf_reader.pages:
        texto += page.extract_text() + "\n"
    return texto

def identificar_modelo(texto):
    if "FILME" in texto.upper():
        return "filme"
    return "saco"

def preencher_planilha(modelo, dados_extraidos):
    if modelo == "saco":
        planilha_path = "SACO.xlsx"
    else:
        planilha_path = "FILME.xlsx"

    wb = load_workbook(planilha_path)
    ws = wb.active

    # Exemplo de preenchimento (adaptar aos campos corretos da planilha)
    campos = {
        "CLIENTE": re.search(r"CLIENTES?:\s*(.*)", dados_extraidos),
        "DATA PEDIDO": re.search(r"DATA PEDIDO:\s*(.*)", dados_extraidos),
        "DATA ENTREGA": re.search(r"DATA DE ENTREGA:\s*(.*)", dados_extraidos),
        "PRODUTO": re.search(r"PRODUTO:\s*(.*)", dados_extraidos),
        "QTDE KG": re.search(r"QTDE. KG:\s*(.*)", dados_extraidos),
        "QTDE MIL": re.search(r"QTDE. \(MIL\):\s*(.*)", dados_extraidos),
        "LARGURA": re.search(r"LARGURA \(mm\):\s*(.*)", dados_extraidos),
        "LARGURA FINAL": re.search(r"LARGURA FINAL \(mm\):\s*(.*)", dados_extraidos),
        "PASSO": re.search(r"PASSO \(mm\):\s*(.*)", dados_extraidos),
        "CILINDRO": re.search(r"CILINDRO \(mm\):\s*(.*)", dados_extraidos),
        "ESPESSURA": re.search(r"ESPESSURA \(p/ parede\):\s*(.*)", dados_extraidos),
        "ESPESSURA FINAL": re.search(r"ESPESSURA FINAL:\s*(.*)", dados_extraidos),
        "OBSERVAÇÕES": re.search(r"OBSERVAÇÕES\n(.*?)\n", dados_extraidos, re.DOTALL),
        "PEDIDO N": re.search(r"PEDIDO N:?\s*(.*)", dados_extraidos),
        "O.C": re.search(r"O.C.:?\s*(.*)", dados_extraidos)
    }

    # Preencher as células da planilha modelo com os dados encontrados
    for row in ws.iter_rows(min_row=2, max_row=2):  # Preenche primeira linha útil
        for cell in row:
            header = cell.column_letter + "1"
            campo_nome = ws[header].value
            if campo_nome in campos and campos[campo_nome]:
                cell.value = campos[campo_nome].group(1).strip()

    # Salvar planilha em memória
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

st.title("Gerador de Fichas Técnicas (SACOS e FILMES)")

uploaded_pdf = st.file_uploader("Envie o PDF da ficha", type=["pdf"])

if uploaded_pdf:
    texto = extrair_dados_pdf(uploaded_pdf)
    modelo = identificar_modelo(texto)
    st.success(f"Modelo detectado: {modelo.upper()}")
    
    planilha_preenchida = preencher_planilha(modelo, texto)
    st.download_button("📥 Baixar planilha preenchida", data=planilha_preenchida, file_name=f"ficha_{modelo}.xlsx")
